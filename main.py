# import cv2
# import face_recognition
# import os
# import re
# import time
# import pickle
# import numpy as np 
# import openpyxl
# from openpyxl import Workbook
# from openpyxl.styles import Font, Alignment, PatternFill


# # Global variables to store detected persons' details
# detected_persons = []
# # def main():
# def record_attendance(name, Employee_ID, attendance_file, status):
#     if os.path.exists(attendance_file):
#         wb = openpyxl.load_workbook(attendance_file)
#     else:
#         wb = Workbook()
#         ws = wb.active
#         heading = ["FOURBRICK TECHNOLOGY SMART CCTV ATTENDANCE SYSTEM"]
#         ws.append(heading)
#         ws.append(["Name", "Date", "Time", 'Employee_ID', 'Status'])

#     current_date = time.strftime("%Y-%m-%d")
#     current_time = time.strftime("%H:%M:%S")

#     ws = wb.active
#     ws.append([name, current_date, current_time, Employee_ID, status])
#     ws.merge_cells('A1:E1')  # Merging cells A1 to E1
#     heading_cell = ws['A1']
#     heading_cell.font = Font(color="FFFFFF", bold=True)
#     heading_cell.fill = PatternFill(start_color="000090", end_color="000080", fill_type="solid")
#     heading_cell.alignment = Alignment(horizontal='center', vertical='center')

#     wb.save(attendance_file)

# def recognize_faces(camera_port, known_faces, known_labels):
#     cam = cv2.VideoCapture(camera_port)

#     id_to_name_mapping = [
#         {'shobhit_\d+': {'name': 'Shobhit Pal', 'Employee_ID': 'FB|IT|CR|008'}},
#         {'vishal_\d+': {'name': 'Vishal Yadav', 'Employee_ID': 'FB|IT|CR|015'}},
#         {'shubhamThakur_\d+': {'name': 'Shubham Thakur', 'Employee_ID': 'FB|IT|CR|024'}},
#         {'sandeepChaudhary_\d+': {'name': 'Sandeep Chaudhary', 'Employee_ID': 'FB|IT|CR|010'}},
#         {'shubham_Chauhan_\d+': {'name': 'Shubham Singh', 'Employee_ID': 'FB|IT|CR|014'}},
#         {'danish_\d+': {'name': 'Danish Khan', 'Employee_ID': 'FB|IT|CR|000'}},
#         {'ashutosh_\d+': {'name': 'Ashutosh Sir', 'Employee_ID': 'FB|IT|CR|000'}},
#         {'nitish_\d+': {'name': 'Nitish Arora', 'Employee_ID': 'FB|IT|CR|028'}},
#         {'mukesh_\d+': {'name': 'Mukesh Kumar', 'Employee_ID': 'FB|IT|CR|005'}},
#         {'deepak_\d+': {'name': 'Deepak Chauhan', 'Employee_ID': 'FB|IT|CR|026'}},
#         {'amrita_\d+': {'name': 'Amrita Singh', 'Employee_ID': 'FB|IT|CR|002'}},
#         {'ravina_\d+': {'name': 'Ravina Singh', 'Employee_ID': 'FB|IT|CR|007'}},
#         {'amit_\d+': {'name': 'Amit Gadiya', 'Employee_ID': 'FB|IT|CR|001'}},
#     ]

#     while True:
#         ret, frame = cam.read()

#         face_locations = face_recognition.face_locations(frame)
#         face_encodings = face_recognition.face_encodings(frame, face_locations)

#         for (top, right, bottom, left), face_encoding in zip(face_locations, face_encodings):
#             matches = face_recognition.compare_faces(known_faces, face_encoding, tolerance=0.5)
#             name = "Unknown"

#             if True in matches:
#                 first_match_index = matches.index(True)
#                 predicted_id = known_labels[first_match_index]

#                 for mapping in id_to_name_mapping:
#                     for pattern, details in mapping.items():
#                         if re.match(pattern, predicted_id):
#                             name = details['name']
#                             Employee_ID = details['Employee_ID']
#                             break

#                 # Check if the person is already detected
#                 if any(person['name'] == name and person['Employee_ID'] == Employee_ID for person in detected_persons):
#                     print(f"Attendance already submitted for {name} (Employee ID: {Employee_ID})")
#                     # Display a message on the frame
#                     cv2.putText(frame, f"{name} Your Attendance already submitted", (left + 6, bottom + 20),
#                                 cv2.FONT_HERSHEY_DUPLEX, 0.5, (255, 255, 255), 1)
#                 else:
#                     # Check the time and set the status
#                     current_time = time.strptime(time.strftime("%H:%M:%S"), "%H:%M:%S")
#                     if current_time < time.strptime("18:30:00", "%H:%M:%S"):
#                         status = "In"
#                     else:
#                         status = "Out"

#                     record_attendance(name, Employee_ID, attendance_file, status)
#                     detected_persons.append({'name': name, 'Employee_ID': Employee_ID})
#                     print(f"Attendance recorded for {name} (Employee ID: {Employee_ID}, Status: {status})")

#                     # Display face detected message on the frame
#                     cv2.putText(frame, "Face Detected", (left + 6, bottom + 20), cv2.FONT_HERSHEY_DUPLEX, 0.5,
#                                 (255, 255, 255), 1)

#         cv2.imshow('Face Recognition', frame)
#         time.sleep(0.1)

#         if cv2.waitKey(1) & 0xFF == ord('q'):
#             break

#     cam.release()
#     cv2.destroyAllWindows()

# if __name__ == "__main__": 
#     camera_port = 0

#     with open("trained_data.pkl", "rb") as f:
#         known_faces, known_labels = pickle.load(f)

#     attendance_file = "attendance.xlsx"

#     recognize_faces(camera_port, known_faces, known_labels)
# # main()






import cv2
import face_recognition
import os
import re
import time
import pickle
import numpy as np 
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
import json
from base import *

is_camera_running = False
detected_persons = []
def main_code():

    def record_attendance_mongo(name,Employee_ID,status):
        attendance_data = {
            'name': name,
            'Employee_ID': Employee_ID,
            'status': status,
            'date': time.strftime("%Y-%m-%d"),
            'time': time.strftime("%H:%M:%S")
        }
        db.employee_data.insert_one(attendance_data)
    # record_attendance_mongo()
    

    def record_attendance(name, Employee_ID, attendance_file, status):
        if os.path.exists(attendance_file):
            wb = openpyxl.load_workbook(attendance_file)
        else:
            wb = Workbook()
            ws = wb.active
            heading = ["FOURBRICK TECHNOLOGY SMART CCTV ATTENDANCE SYSTEM"]
            ws.append(heading)
            ws.append(["Name", "Date", "Time", 'Employee_ID', 'Status'])

        current_date = time.strftime("%Y-%m-%d")
        current_time = time.strftime("%H:%M:%S")

        ws = wb.active
        ws.append([name, current_date, current_time, Employee_ID, status])
        ws.merge_cells('A1:E1')  # Merging cells A1 to E1
        heading_cell = ws['A1']
        heading_cell.font = Font(color="FFFFFF", bold=True)
        heading_cell.fill = PatternFill(start_color="000090", end_color="000080", fill_type="solid")
        heading_cell.alignment = Alignment(horizontal='center', vertical='center')

        wb.save(attendance_file)

    def load_employee_data(file_path):
        with open(file_path, 'r') as file:
            data = json.load(file)
        return data

    def recognize_faces(camera_port, known_faces, known_labels, employee_data):
        cam = cv2.VideoCapture(camera_port)

        while True:
            ret, frame = cam.read()

            face_locations = face_recognition.face_locations(frame)
            face_encodings = face_recognition.face_encodings(frame, face_locations)

            for (top, right, bottom, left), face_encoding in zip(face_locations, face_encodings):
                matches = face_recognition.compare_faces(known_faces, face_encoding, tolerance=0.5)
                name = "Unknown"

                if True in matches:
                    first_match_index = matches.index(True)
                    predicted_id = known_labels[first_match_index]

                    for details in employee_data:
                        for pattern, employee_info in details.items():
                            if re.match(pattern, predicted_id):
                                name = employee_info['name']
                                Employee_ID = employee_info['Employee_ID']
                                break

                    if any(person['name'] == name and person['Employee_ID'] == Employee_ID for person in detected_persons):
                        print(f"Attendance already submitted for {name} (Employee ID: {Employee_ID})")
                        cv2.putText(frame, f"{name} Your Attendance already submitted", (left + 6, bottom + 20),
                                    cv2.FONT_HERSHEY_DUPLEX, 0.5, (255, 255, 255), 1)
                    else:
                        current_time = time.strptime(time.strftime("%H:%M:%S"), "%H:%M:%S")
                        status = "In" if current_time < time.strptime("18:30:00", "%H:%M:%S") else "Out"
                        
                        record_attendance_mongo(name, Employee_ID, status)

                        record_attendance(name, Employee_ID, attendance_file, status)
                        detected_persons.append({'name': name, 'Employee_ID': Employee_ID})
                        print(f"Attendance recorded for {name} (Employee ID: {Employee_ID}, Status: {status})")
                        cv2.putText(frame, "Face Detected", (left + 6, bottom + 20), cv2.FONT_HERSHEY_DUPLEX, 0.5,
                                    (255, 255, 255), 1)

            cv2.imshow('Face Recognition', frame)
            time.sleep(0.1)

            if cv2.waitKey(1) & 0xFF == ord('q'):
                break

        cam.release()
        cv2.destroyAllWindows()
    attendance_file ="attendance.xlsx"
    def mmm():
        # global attendance_file
        camera_port = 0

        with open("trained_data.pkl", "rb") as f:
            known_faces, known_labels = pickle.load(f)

        # attendance_file = "attendance.xlsx"
        employee_data_file = "employee_data.json"

        id_to_name_mapping = load_employee_data(employee_data_file)

        recognize_faces(camera_port, known_faces, known_labels, id_to_name_mapping)
    mmm()

    


# main_code()




